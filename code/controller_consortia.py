try:
	
	import win32com.client as win32
	import pythoncom
	import calendar
	
	import xlwings as xw

	import pandas as pd
	import numpy as np
	import csv, os , glob

	from pathlib import Path
	from datetime import datetime, date

	# Importo toda la informacion de mi controllador
	from controller_consortia import *


	from openpyxl import load_workbook
	from openpyxl import Workbook
    # Esto libreria nos ayuda a crear pivot tables
	from openpyxl.utils.dataframe import dataframe_to_rows
	from openpyxl.utils import get_column_letter
        #------------- Tablas e imagenes -------------------------
    # Primero importo algunas funciones para tablas e imagenes
	from openpyxl.worksheet.table import Table, TableStyleInfo

	from openpyxl.drawing.image import Image


	
	win32c = win32.constants

except ModuleNotFoundError as err:
    print('Opssss... Looks like there is an error importing the package', err)

class ReportC():
    #------------- CONSTRUCTOR ------------------
    # Inicializo el constructor
    # Cada vez que cree un objeto de tipo report se inicializan estos atributos
    # Se repiten en todos los metodos aunque cambia el valores# El valor lo cambio directo desde el archivo de rutas
    def __init__(self, path):
        
        #self.type = type

        print('Este es un path--------', path)

        #----------- Inicializo objeto de Windows
        self.Excel = win32.gencache.EnsureDispatch('Excel.Application',pythoncom.CoInitialize())


        #--------- LISTA DE DATAFRAMES
        datasets = []  

        for file_path in path:

            #route_path = str("{}.xlsx".format(file_path))

            print ('.............................',file_path)

            df = pd.read_excel(file_path)

            print('El DF se pudo crear')
            datasets.append(df)
            print('El data set se pudo adjuntat al array')

        # Access the individual datasets
        #for i, dataset in enumerate(datasets):

        #-------- MERGE the data into a single dataset.

        df = pd.concat(datasets)


        #---------- LIMPIO DATA
        # Vamos a quitar todos los campos que no tienen SESSION & TOTAL ITEM INVESTIGATIONS

        #---------- LIMPIO DATA
        # Vamos a quitar todos los campos que no tienen SESSION & TOTAL ITEM INVESTIGATIONS

        # Si ambas columnas no tienen datos creo un nuevo data set con la informacion que si tiene
        df = df[(df['Sessions'] != 0) & (df['Total_Item_Investigations'] != 0)]

        #----------- SAVE ALL FILE
        #Aqui guardo la data ya trabajada


        df.to_excel('filtered_data.xlsx', index=False)

        print('----------------FILTERED DATA hava been created susscessfully----------------')

        self.new_column()
        self.save()
        self.clean_folder() 
	#---------- CREATE NEW COLUMN -MONTH-YEAR- --------------
    def save(self):
        #------------ ABRO EL ORIGEN Y EL DESTINO 
        # -----ORIGEN
        wb = load_workbook('final_path.xlsx')
        ws = wb.worksheets[0]

        #---- DESTINO  ------------------

        #file_path = "C:/wamp_latest/www/PYTHON/Templates/webstats/controller_template.xlsm"
        # Construct the relative path to the file
        file_path = os.path.join(os.getcwd(), "consortia_controller_template.xlsm")


        # Primero abro el libro que ya tiene la columna incluida
        wb1 = self.Excel.Workbooks.Open(file_path)

        # Esta es la forma de seleccionar el SHEET dentro de work book.
        ws1 = wb1.Worksheets(1)  # Activate the first worksheet in the workbook


        # Esto es para meterle informcion adicional a los encabezados
        # Como el titulo y demas.
        ws2 = wb1.Worksheets(3)



        # Averiguo el tamaño del archivo original que deseamos COPIAR
        start_row = 2
        max_row = ws.max_row

        start_col = 1
        max_col = ws.max_column


        # Get the source data in a nested list
        source_data = []

        #------------ COPIO INFO

        for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=start_col, max_col=max_col, values_only=True):
            source_data.append(list(row))


        #------------ PASTE INFO
        # Paste the copied data into the target worksheet

        # Calculate the target range
        num_rows = len(source_data)
        num_cols = len(source_data[0])


        # Escojo el mobre de la tabla para pegar la info
        target_table = ws1.ListObjects("Table2")  

        # Get the range of the target table
        table_range = target_table.Range


        # Calculate the dimensions of the source data
        num_rows = len(source_data)
        num_cols = len(source_data[0])




        # Calculate the target range within the table
        table_start_row = table_range.Row + table_range.Rows.Count
        table_start_col = table_range.Column
        table_end_row = table_start_row + num_rows - 1
        table_end_col = table_start_col + num_cols - 1


        #---------- DELETE CONTENT
        # Genero el MAX row de ws1 para poder limpiar la tabla
        # para ello genero un table range adicional con la informacion que deseo borrar.


        # Find the last used row in column A
        last_row = ws1.Cells(ws1.Rows.Count, 1).End(win32.constants.xlUp).Row

        # Defino el rango que deseo eliminar
        table_range1 = ws1.Range(
            ws1.Cells(start_row, start_col),
            ws1.Cells(last_row, max_col)
        )

        # Aplico cliear contents al rango
        #No se le pueden pasar parametros tiene que ser asi
        table_range1.ClearContents()

        print('Table cleared. ready for inyection of new data...')

        #------------ TRANFER INFORMATION
        # Define the target range within the table
        target_range = ws1.Range(
            ws1.Cells(start_row, start_col),
            ws1.Cells(max_row, max_col)
        )


        # Assign the source data to the target range
        target_range.Value = source_data
        print('----------------')
        print('Data has been uploaded properly ...')

        print('Waiting for the autofilled option in sheet 3 ...')


        #----------------- Titulo del reporte
        # Esto es una prueba para comprobar el access a la hoja3 con los graficos
        # Lo vamosa utilizar para generar el titulo igual que el
        ws2.Range("C5").Value = 'Webstats report generated for: '
        ws2.Range("C5").Font.Size = 20
        ws2.Range("C5").Font.Bold = True

        ws2.Range("C7").Value = 'Type:'
        ws2.Range("C7").Font.Size = 20
        ws2.Range("C7").Font.Bold = True

        ws2.Range("C9").Value = 'Period analyzed:'
        ws2.Range("C9").Font.Size = 20
        ws2.Range("C9").Font.Bold = True


        #---type
        ws2.Range("I7").Value = str("{} report".format('Consortia'))
        ws2.Range("I7").Font.Size = 28
        ws2.Range("I7").Font.Bold = True
        ws2.Range("I7").Font.Color = self.RGB(128,128,128)

        #-----Name
        ws2.Range("I5").Value = str('Consortia group')
        ws2.Range("I5").Font.Size = 28
        ws2.Range("I5").Font.Bold = True
        ws2.Range("I5").Font.Color = self.RGB(128,128,128)

        #----- Date range
        #ws2.Range("I9").Value = str('{} {} - {} {}'.format(self.initMonth,self.startYear, self.endMonth, self.closeYear))
        #ws2.Range("I9").Font.Size = 28
        #ws2.Range("I9").Font.Bold = True
        #ws2.Range("I9").Font.Color = self.RGB(128,128,128)


        print('-----------------')
        print('Data has been generated ...')

        # CLose True hace los mismo que save pero cierra el libro
        wb1.Close(True)
        

        print('Closing COM')
        # Tengo que cerrar el objeto de excel porque sino no me deja abrir el spreadsheet
        self.Excel.Quit()

    def load(self):
  
        xw.Book('consortia_controller_template.xlsm')

    def new_column(self):
        #-----------------
        #  OPEN FILE
        #-----------------
        print('NEW COLUMN()-----------')
        print('---- Open Filtered data file ----')
        # Ahora vamos a buscar el archivo y cargarlo
        wb = load_workbook('filtered_data.xlsx')
        ws = wb.active

        # Ahora creamos la cell que queremos agregar
        # Hay que tener en cuanta el tamaño del sheet para poder ubicarla bien en la columna correcta
        # Creo una variable donde guardo la cell

        newCol = ws['AE1']
        newCol.value = 'Month-Year'

        #-------------------
        #  FIND SIZE DF
        #------------------
        # Podemos averiguar el tamaño de la coumna facilment con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel('filtered_data.xlsx', engine='openpyxl')

        rows = df.shape[0]
        cols = df.shape[1]
        # Esto es para poder saber en donde tengo que poner la ultima columna en caso de que cambie el numero de columnas
        size = df.shape[0]
        #-------------------------
        #  FOR LOOP NEW COLUMN
        #--------------------------
        # Ahora creamos un forloop para interactuar con todos los rows en estas columnas
        # RECORDAR que no debemos incluir el primer row porque son HEADING y no values
        # Ponemos de limite la variable 'df'
        # Le ponemos el size mas 2 porque me fglatan dos espacio
        # Le decimos que comience desde 2 pero el DF me cuenta solo los spacios con data y los los encabezados
        # Para compensar esto hacemos el truco

        for row in range(2,(size+2)):
            # En cada interaccion pillamos el numero de row para el MES y el año
            # Lo convertimos a STR para poder hacerlo una cadena de texto
            m = str(ws[f'AC{row}'].value)
            y = str(ws[f'AD{row}'].value)


            # Esto es para verificar si los valroes que me llegan son validos
            if m is not None and y is not None:
                m = str(m)
                y = str(y)
                # Creamos la cadena de texto con la fecha complete
                a_date = "1"+"/"+m+"/"+y

                # En el último parametro especifico el formato de fecha que deseamos tener '%B %Y'
                # B% da el mes con nombre completo
                # b% nos da la abreviatura del mes
                # m% nos da el numero del mes

                #------------------------------
                #       NEW columna
                #------------------------------
                # En la columna AE ya le hemos asignado nombre
                # Ahora necesitamos asignarle los valores
                ws[f'AE{row}'] = datetime.strptime(a_date, "%d/%m/%Y").strftime('%b-%Y')

        print('---- File Columnd has been created ----')
        #Salvamos el doc como excel
        wb.save('column.xlsx')


        df1 = pd.read_excel('column.xlsx')

        # select  columns to display
        dframe = df1[['Title',
                'Unique_Item_Requests',
                'Total_Item_Requests',
                'Total_Item_Investigations',
                'Unique_Item_Investigations',
                'Sessions',
                'Platform',
                'Subject',
                'OrderDescription',
                'OrderNumber',
                'UsedByCustomer',
                'Group',
                'User',
                'Month',
                'Year',
                'Month-Year']]

        #---------------------
        #  FIND SIZE NEW DF
        #---------------------
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        newRows = dframe.shape[0]
        newCols = dframe.shape[1]

        #dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime('%b-%Y')
        #dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime("%Y-%m-%d")

        dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime("%b-%Y")

        # A esta tabla le quitamos el formato de tabla para incluir la ultima columna que creamos con la fecha
        # Le ponemos el index flase para que no aparezca la coumna index
        dframe.to_excel('final_path.xlsx', index = False)
        print('---- FinalPath file has been created ----')

    def clean_folder(self):

        # Esta fucnion remoeve todos los archivos de excel del directorio.x
        # Somo me quedo con los templates que estan en MACRO
        os.getcwd()

        files = glob.glob("*.xlsx")

        for file in files:
            os.remove(file)
        print("xlsx files removed successfully.")

    def RGB(self,r,g,b):

        bgr = (r, g, b)
        strValue = '%02x%02x%02x' % bgr
        # print(strValue)
        iValue = int(strValue, 16)
        return iValue
    
